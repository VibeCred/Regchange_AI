"""
RegChange AI — Main FastAPI Application
API server for the regulatory change intelligence platform.
"""
import os
import sys
import uuid
import json
import shutil
import logging
import asyncio
import traceback
from datetime import datetime, timezone
from typing import Optional

from fastapi import FastAPI, UploadFile, File, HTTPException, Query, BackgroundTasks, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel
import io

# Add parent to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from backend.config import (
    UPLOAD_DIR, DATA_DIR, SUPPORTED_EXTENSIONS, MAX_UPLOAD_SIZE_MB,
    API_PREFIX, CHANGE_CATEGORIES
)
from backend.database.db import Database
from backend.pipeline.pdf_extractor import PDFExtractor
from backend.pipeline.normalizer import TextNormalizer
from backend.pipeline.structure_parser import StructureParser
from backend.pipeline.clause_aligner import ClauseAligner, AlignmentType
from backend.pipeline.change_classifier import ChangeClassifier
from backend.pipeline.impact_scorer import ImpactScorer
from backend.pipeline.confidence_engine import ConfidenceEngine
from backend.pipeline.llm_classifier import LLMClassifier
from backend.models.document import DocumentVersion, ParsedDocument
from backend.models.change import ChangeType, ComparisonResult

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s'
)
logger = logging.getLogger("regchange")

# Initialize app
app = FastAPI(
    title="RegChange AI",
    description="AI-Powered Regulatory Change Intelligence Platform",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize components
db = Database()
pdf_extractor = PDFExtractor()
normalizer = TextNormalizer()
structure_parser = StructureParser()
change_classifier = ChangeClassifier()
impact_scorer = ImpactScorer()
confidence_engine = ConfidenceEngine()

# Global state for progress tracking
comparison_progress = {}


# === Request/Response Models ===

class CompareRequest(BaseModel):
    old_document_id: str
    new_document_id: str


class ReviewRequest(BaseModel):
    status: str  # ACCEPTED, REJECTED, EDITED, FLAGGED
    comment: str = ""


# === API Endpoints ===

@app.post(f"{API_PREFIX}/documents/upload")
async def upload_document(
    file: UploadFile = File(...),
    version: str = Query("old", regex="^(old|new)$"),
):
    """Upload a PDF document."""
    # Validate file
    filename = file.filename or "unknown.pdf"
    ext = os.path.splitext(filename)[1].lower()
    
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(400, f"Unsupported file type: {ext}")
    
    # Read and save file
    content = await file.read()
    
    if len(content) > MAX_UPLOAD_SIZE_MB * 1024 * 1024:
        raise HTTPException(400, f"File too large. Max: {MAX_UPLOAD_SIZE_MB}MB")
    
    document_id = f"DOC_{version}_{uuid.uuid4().hex[:8]}"
    file_path = os.path.join(UPLOAD_DIR, f"{document_id}{ext}")
    
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Extract metadata
    try:
        metadata = pdf_extractor.extract_metadata(file_path)
        extracted = pdf_extractor.extract(file_path)
        
        doc_data = {
            "document_id": document_id,
            "filename": filename,
            "title": metadata.get("title", ""),
            "circular_number": metadata.get("circular_number", ""),
            "issue_date": metadata.get("issue_date", ""),
            "total_pages": extracted["total_pages"],
            "version": version,
            "quality_score": extracted["quality_score"],
            "file_path": file_path,
        }
        
        db.save_document(doc_data)
        
        return {
            "document_id": document_id,
            "filename": filename,
            "total_pages": extracted["total_pages"],
            "quality_score": round(extracted["quality_score"], 3),
            "title": metadata.get("title", ""),
            "circular_number": metadata.get("circular_number", ""),
            "version": version,
        }
    
    except Exception as e:
        # Clean up on failure
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(500, f"Failed to process document: {str(e)}")


@app.post(f"{API_PREFIX}/comparisons")
async def start_comparison(
    request: CompareRequest,
    background_tasks: BackgroundTasks,
):
    """Start a comparison between two documents."""
    comparison_id = f"CMP_{uuid.uuid4().hex[:8]}"
    
    # Save initial comparison record
    comp_data = {
        "comparison_id": comparison_id,
        "old_document_id": request.old_document_id,
        "new_document_id": request.new_document_id,
        "status": "processing",
        "progress": 0.0,
        "current_stage": "Initializing...",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    db.save_comparison(comp_data)
    
    # Initialize progress tracking
    comparison_progress[comparison_id] = {
        "progress": 0.0,
        "stage": "Initializing...",
        "status": "processing",
    }
    
    # Run comparison in background
    background_tasks.add_task(run_comparison_pipeline, comparison_id, 
                              request.old_document_id, request.new_document_id)
    
    return {
        "comparison_id": comparison_id,
        "status": "processing",
        "message": "Comparison started. Poll /comparisons/{id} for progress.",
    }


@app.get(f"{API_PREFIX}/comparisons/{{comparison_id}}")
async def get_comparison(comparison_id: str):
    """Get comparison status and results."""
    # Check in-memory progress first
    if comparison_id in comparison_progress:
        progress = comparison_progress[comparison_id]
        comp = db.get_comparison(comparison_id)
        if comp:
            comp['progress'] = progress['progress']
            comp['current_stage'] = progress['stage']
            comp['status'] = progress['status']
            
            # Parse statistics JSON if present
            if isinstance(comp.get('statistics'), str):
                try:
                    comp['statistics'] = json.loads(comp['statistics'])
                except:
                    comp['statistics'] = {}
            
            return comp
    
    comp = db.get_comparison(comparison_id)
    if not comp:
        raise HTTPException(404, "Comparison not found")
    
    if isinstance(comp.get('statistics'), str):
        try:
            comp['statistics'] = json.loads(comp['statistics'])
        except:
            comp['statistics'] = {}
    
    return comp


@app.get(f"{API_PREFIX}/comparisons/{{comparison_id}}/changes")
async def get_changes(
    comparison_id: str,
    category: Optional[str] = None,
    impact: Optional[str] = None,
    change_type: Optional[str] = None,
    is_substantive: Optional[bool] = None,
    review_status: Optional[str] = None,
    search: Optional[str] = None,
):
    """Get changes for a comparison with optional filters."""
    filters = {}
    if category:
        filters['category'] = category
    if impact:
        filters['impact'] = impact
    if change_type:
        filters['change_type'] = change_type
    if is_substantive is not None:
        filters['is_substantive'] = is_substantive
    if review_status:
        filters['review_status'] = review_status
    
    changes = db.get_changes(comparison_id, filters)
    
    # Apply text search if provided
    if search:
        search_lower = search.lower()
        changes = [
            c for c in changes
            if search_lower in c.get('change_summary', '').lower()
            or search_lower in c.get('old_requirement', '').lower()
            or search_lower in c.get('new_requirement', '').lower()
        ]
    
    return {"changes": changes, "total": len(changes)}


@app.get(f"{API_PREFIX}/comparisons/{{comparison_id}}/changes/{{change_id}}")
async def get_change_detail(comparison_id: str, change_id: str):
    """Get detailed change information."""
    change = db.get_change(change_id)
    if not change:
        raise HTTPException(404, "Change not found")
    return change


@app.get(f"{API_PREFIX}/comparisons/{{comparison_id}}/statistics")
async def get_statistics(comparison_id: str):
    """Get comparison statistics."""
    stats = db.get_statistics(comparison_id)
    if not stats:
        raise HTTPException(404, "Comparison not found")
    return stats


@app.get(f"{API_PREFIX}/comparisons/{{comparison_id}}/export/excel")
async def export_excel(comparison_id: str):
    """Export all detected regulatory changes as an Excel workbook (.xlsx)."""
    comp = db.get_comparison(comparison_id)
    if not comp:
        raise HTTPException(404, "Comparison not found")
    
    changes = db.get_changes(comparison_id)
    if not changes:
        raise HTTPException(404, "No changes found for this comparison")
    
    # Build rows
    rows = _build_export_rows(changes)
    
    # Generate Excel in memory using openpyxl directly (faster than pandas)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Regulatory Changes"
        
        # Headers
        headers = list(rows[0].keys()) if rows else []
        ws.append(headers)
        
        # Header styling
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data rows
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        
        # Column widths
        col_widths = [14, 16, 14, 15, 25, 12, 12, 40, 20, 45, 20, 45, 30, 30, 35, 15, 25]
        for i, width in enumerate(col_widths):
            if i < len(headers):
                ws.column_dimensions[chr(65 + i)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"RegChange_AI_{comparison_id}_Changes.xlsx"
        resp_headers = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=resp_headers,
        )
    except Exception as e:
        logger.error(f"Excel export failed: {e}\n{traceback.format_exc()}")
        raise HTTPException(500, f"Failed to generate Excel export: {str(e)}")


@app.get(f"{API_PREFIX}/comparisons/{{comparison_id}}/export/csv")
async def export_csv(comparison_id: str):
    """Export all detected regulatory changes as a CSV file."""
    comp = db.get_comparison(comparison_id)
    if not comp:
        raise HTTPException(404, "Comparison not found")
    
    changes = db.get_changes(comparison_id)
    if not changes:
        raise HTTPException(404, "No changes found for this comparison")
    
    rows = _build_export_rows(changes)
    
    import csv as csv_module
    output = io.StringIO()
    if rows:
        writer = csv_module.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    
    csv_bytes = output.getvalue().encode('utf-8-sig')
    
    filename = f"RegChange_AI_{comparison_id}_Changes.csv"
    resp_headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Access-Control-Expose-Headers": "Content-Disposition"
    }
    
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers=resp_headers,
    )


def _build_export_rows(changes):
    """Build structured rows for export from change records."""
    rows = []
    for c in changes:
        old_ref = c.get('old_reference') or {}
        new_ref = c.get('new_reference') or {}
        
        # Parse JSON if stored as strings
        if isinstance(old_ref, str):
            try:
                old_ref = json.loads(old_ref)
            except:
                old_ref = {}
        if isinstance(new_ref, str):
            try:
                new_ref = json.loads(new_ref)
            except:
                new_ref = {}
        
        old_loc = f"Page {old_ref.get('page', 'N/A')}"
        if old_ref.get('section') or old_ref.get('clause'):
            old_loc += f" ({old_ref.get('section') or old_ref.get('clause')})"
            
        new_loc = f"Page {new_ref.get('page', 'N/A')}"
        if new_ref.get('section') or new_ref.get('clause'):
            new_loc += f" ({new_ref.get('section') or new_ref.get('clause')})"
            
        # Numerical changes description
        num_changes = c.get('numerical_changes') or []
        if isinstance(num_changes, str):
            try:
                num_changes = json.loads(num_changes)
            except:
                num_changes = []
        num_desc_parts = []
        for nc in num_changes:
            if isinstance(nc, dict):
                part = f"{nc.get('field', 'value').title()}: {nc.get('old_value')} -> {nc.get('new_value')} ({nc.get('direction', '')}"
                if nc.get('magnitude_percent'):
                    part += f", {nc.get('magnitude_percent'):.0f}%"
                part += ")"
                num_desc_parts.append(part)
        num_desc = "; ".join(num_desc_parts)
        
        # Obligation change description
        ob_change = c.get('obligation_change') or {}
        if isinstance(ob_change, str):
            try:
                ob_change = json.loads(ob_change)
            except:
                ob_change = {}
        ob_desc = ""
        if isinstance(ob_change, dict) and ob_change.get('direction') and ob_change.get('direction') != 'UNCHANGED':
            ob_desc = f"{ob_change.get('direction')}: {ob_change.get('explanation', '')}"
        
        cat_code = c.get('category', '')
        cat_name = CHANGE_CATEGORIES.get(cat_code, cat_code)
        
        rows.append({
            "Change ID": c.get('change_id', ''),
            "Impact Level": c.get('impact', 'INFORMATIONAL'),
            "Change Type": c.get('change_type', ''),
            "Category Code": cat_code,
            "Category Name": cat_name,
            "Substantive": "Yes" if c.get('is_substantive') else "No",
            "Confidence": f"{round((c.get('confidence_overall', 0) or 0) * 100)}%",
            "Change Summary": c.get('change_summary', ''),
            "Old Location": old_loc,
            "Old Requirement Text": c.get('old_requirement', ''),
            "New Location": new_loc,
            "New Requirement Text": c.get('new_requirement', ''),
            "Numerical Changes": num_desc,
            "Obligation Analysis": ob_desc,
            "AI Impact Explanation": c.get('llm_explanation') or c.get('impact_explanation', ''),
            "Review Status": c.get('review_status', 'PENDING'),
            "Reviewer Comment": c.get('reviewer_comment', ''),
        })
    
    return rows


@app.post(f"{API_PREFIX}/changes/{{change_id}}/review")
async def review_change(change_id: str, request: ReviewRequest):
    """Submit a review for a change."""
    change = db.get_change(change_id)
    if not change:
        raise HTTPException(404, "Change not found")
    
    db.update_review(change_id, request.status, request.comment)
    return {"status": "ok", "change_id": change_id, "review_status": request.status}


@app.get(f"{API_PREFIX}/comparisons")
async def list_comparisons():
    """List all comparisons."""
    return {"comparisons": db.list_comparisons()}


@app.get(f"{API_PREFIX}/categories")
async def get_categories():
    """Get change category definitions."""
    return {"categories": CHANGE_CATEGORIES}


# === Comparison Pipeline ===

async def run_comparison_pipeline(
    comparison_id: str,
    old_doc_id: str,
    new_doc_id: str,
):
    """Run the full comparison pipeline."""
    try:
        def update_progress(stage: str, progress: float, message: str = ""):
            comparison_progress[comparison_id] = {
                "progress": progress,
                "stage": message or stage,
                "status": "processing",
            }
            db.update_comparison_progress(comparison_id, progress, message or stage)
        
        # Get document file paths
        old_doc = db.get_comparison(comparison_id)
        
        # Find document files
        old_file = _find_document_file(old_doc_id)
        new_file = _find_document_file(new_doc_id)
        
        if not old_file or not new_file:
            raise ValueError("Document files not found")
        
        # === PHASE 1: PDF Extraction ===
        update_progress("extraction", 0.05, "Extracting text from old document...")
        old_extracted = pdf_extractor.extract(old_file)
        
        update_progress("extraction", 0.10, "Extracting text from new document...")
        new_extracted = pdf_extractor.extract(new_file)
        
        # === PHASE 2: Structure Parsing ===
        update_progress("parsing", 0.15, "Parsing old document structure...")
        old_parsed = structure_parser.parse(old_extracted, DocumentVersion.OLD, old_doc_id)
        
        update_progress("parsing", 0.25, "Parsing new document structure...")
        new_parsed = structure_parser.parse(new_extracted, DocumentVersion.NEW, new_doc_id)
        
        logger.info(
            f"Old doc: {len(old_parsed.nodes)} nodes, {len(old_parsed.get_clauses())} clauses. "
            f"New doc: {len(new_parsed.nodes)} nodes, {len(new_parsed.get_clauses())} clauses."
        )
        
        # === PHASE 3: Clause Alignment ===
        update_progress("alignment", 0.30, "Aligning clauses between documents...")
        
        try:
            aligner = ClauseAligner(use_semantic=True)
        except Exception:
            logger.warning("Semantic matching unavailable, using lexical only")
            aligner = ClauseAligner(use_semantic=False)
        
        alignments = aligner.align(old_parsed, new_parsed, progress_callback=None)
        
        update_progress("alignment", 0.50, f"Aligned {len(alignments)} clause pairs")
        
        # === PHASE 4: Change Detection & Classification ===
        update_progress("classification", 0.55, "Detecting and classifying changes...")
        
        change_records = []
        change_counter = 0
        
        for i, alignment in enumerate(alignments):
            # Skip unchanged pairs
            if alignment.alignment_type not in (AlignmentType.ADDED, AlignmentType.REMOVED):
                if alignment.old_nodes and alignment.new_nodes:
                    old_text = alignment.old_nodes[0].normalized_text
                    new_text = alignment.new_nodes[0].normalized_text
                    if old_text == new_text:
                        continue
            
            change_counter += 1
            change_id = f"CHG-{change_counter:04d}"
            
            record = change_classifier.classify_alignment(
                alignment, change_id, comparison_id
            )
            
            # Skip UNCHANGED type
            if record.change_type == ChangeType.UNCHANGED:
                change_counter -= 1
                continue
            
            change_records.append(record)
            
            if i % 20 == 0:
                progress = 0.55 + (i / max(len(alignments), 1)) * 0.15
                update_progress("classification", progress,
                              f"Classifying change {i+1}/{len(alignments)}...")
        
        # === PHASE 5: Impact Scoring ===
        update_progress("impact", 0.72, "Scoring impact levels...")
        change_records = impact_scorer.score_all(change_records)
        
        # === PHASE 6: Confidence Calculation ===
        update_progress("confidence", 0.78, "Computing confidence scores...")
        change_records = confidence_engine.compute_all(change_records)
        
        # === PHASE 7: Optional LLM Enhancement ===
        update_progress("llm", 0.82, "Checking LLM availability...")
        
        llm_available = False
        try:
            llm = LLMClassifier()
            if llm.is_available():
                llm_available = True
                update_progress("llm", 0.83, "Enhancing changes with LLM...")
                
                # Only enhance high-impact substantive changes
                high_impact = [c for c in change_records 
                             if c.is_substantive and c.impact.value in ('CRITICAL', 'HIGH')]
                
                for idx, change in enumerate(high_impact):
                    llm.enhance_change(change)
                    if idx % 5 == 0:
                        progress = 0.83 + (idx / max(len(high_impact), 1)) * 0.10
                        update_progress("llm", progress,
                                      f"LLM processing {idx+1}/{len(high_impact)}...")
            else:
                update_progress("llm", 0.93, "LLM not available — using deterministic results")
        except Exception as e:
            logger.warning(f"LLM enhancement skipped: {e}")
            update_progress("llm", 0.93, "LLM enhancement skipped")
        
        # === PHASE 8: Save Results ===
        update_progress("saving", 0.95, "Saving results...")
        
        # Convert to dicts for storage
        change_dicts = []
        for record in change_records:
            d = {
                "change_id": record.change_id,
                "comparison_id": comparison_id,
                "change_type": record.change_type.value,
                "category": record.category.value,
                "sub_category": record.sub_category,
                "is_substantive": record.is_substantive,
                "impact": record.impact.value,
                "confidence_overall": record.confidence.overall,
                "confidence_breakdown": record.confidence.model_dump(),
                "old_reference": record.old_reference.model_dump() if record.old_reference else {},
                "new_reference": record.new_reference.model_dump() if record.new_reference else {},
                "change_summary": record.change_summary,
                "old_requirement": record.old_requirement,
                "new_requirement": record.new_requirement,
                "impact_explanation": record.impact_explanation,
                "numerical_changes": [nc.model_dump() for nc in record.numerical_changes],
                "obligation_change": record.obligation_change.model_dump() if record.obligation_change else {},
                "evidence": record.evidence,
                "diff_highlights": record.diff_highlights,
                "llm_classification": record.llm_classification or {},
                "llm_explanation": record.llm_explanation,
                "llm_available": record.llm_available,
                "prompt_version": record.prompt_version,
                "model_version": record.model_version,
                "review_status": record.review_status.value,
            }
            change_dicts.append(d)
        
        db.save_changes(change_dicts)
        
        # Update comparison statistics
        stats = db.get_statistics(comparison_id)
        
        comp_update = {
            "comparison_id": comparison_id,
            "old_document_id": old_doc_id,
            "new_document_id": new_doc_id,
            "status": "completed",
            "progress": 1.0,
            "current_stage": "Analysis complete",
            "total_changes": stats.get('total', 0),
            "substantive_changes": stats.get('substantive', 0),
            "editorial_changes": stats.get('editorial', 0),
            "statistics": stats,
            "llm_available": llm_available,
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        db.save_comparison(comp_update)
        
        comparison_progress[comparison_id] = {
            "progress": 1.0,
            "stage": "Analysis complete",
            "status": "completed",
        }
        
        logger.info(
            f"Comparison {comparison_id} complete: "
            f"{len(change_records)} changes detected "
            f"({stats.get('substantive', 0)} substantive, "
            f"{stats.get('editorial', 0)} editorial)"
        )
    
    except Exception as e:
        logger.error(f"Comparison failed: {e}\n{traceback.format_exc()}")
        
        comparison_progress[comparison_id] = {
            "progress": 0.0,
            "stage": f"Error: {str(e)}",
            "status": "failed",
        }
        
        comp_update = {
            "comparison_id": comparison_id,
            "old_document_id": old_doc_id,
            "new_document_id": new_doc_id,
            "status": "failed",
            "error_message": str(e),
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        db.save_comparison(comp_update)


def _find_document_file(document_id: str) -> Optional[str]:
    """Find the file path for a document."""
    conn = db._get_conn()
    row = conn.execute(
        "SELECT file_path FROM documents WHERE document_id = ?", (document_id,)
    ).fetchone()
    conn.close()
    
    if row and row['file_path'] and os.path.exists(row['file_path']):
        return row['file_path']
    
    # Try to find in uploads directory
    for f in os.listdir(UPLOAD_DIR):
        if f.startswith(document_id):
            return os.path.join(UPLOAD_DIR, f)
    
    return None


# Serve frontend static assets
FRONTEND_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "frontend")

@app.get("/")
async def serve_frontend():
    index_path = os.path.join(FRONTEND_DIR, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return {"message": "RegChange AI API is running. Frontend not found."}


if os.path.exists(FRONTEND_DIR):
    app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")
    app.mount("/frontend", StaticFiles(directory=FRONTEND_DIR), name="frontend")
    app.mount("/assets", StaticFiles(directory=FRONTEND_DIR), name="assets")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
